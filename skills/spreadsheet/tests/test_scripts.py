"""Tests for spreadsheet skill Python scripts."""

import json
import subprocess
import sys
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import pytest
import openpyxl
import pandas
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.worksheet.views import Selection

SCRIPTS = Path(__file__).parent.parent / "scripts"


def run(script: str, *args: str) -> subprocess.CompletedProcess:
    return subprocess.run(
        [sys.executable, str(SCRIPTS / script), *args],
        capture_output=True,
        text=True,
    )


@pytest.fixture(scope="session")
def sample_xlsx(tmp_path_factory) -> Path:
    path = tmp_path_factory.mktemp("xlsx") / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["Name", "Age", "Score"])
    ws.append(["Alice", 30, 95])
    ws.append(["Bob", 25, 82])
    ws.append(["Carol", 35, 91])
    wb.save(str(path))
    return path


@pytest.fixture(scope="session")
def sample_csv(tmp_path_factory) -> Path:
    path = tmp_path_factory.mktemp("csv") / "sample.csv"
    path.write_text("Name,Age,Score\nAlice,30,95\nBob,25,82\nCarol,35,91\n")
    return path


@pytest.fixture(scope="session")
def sample_parquet(tmp_path_factory) -> Path:
    """The same rows as the other fixtures, plus a dictionary-encoded column."""
    path = tmp_path_factory.mktemp("parquet") / "sample.parquet"
    frame = pandas.DataFrame(
        {
            "Name": ["Alice", "Bob", "Carol"],
            "Age": [30, 25, 35],
            "Score": [95, 82, 91],
            "Region": pandas.Categorical(["East", "West", "East"]),
        }
    )
    frame.to_parquet(path, index=False)
    return path


@pytest.fixture(scope="session")
def large_parquet(tmp_path_factory) -> Path:
    """More rows than a default preview shows, to exercise batched reads."""
    path = tmp_path_factory.mktemp("parquet-large") / "large.parquet"
    frame = pandas.DataFrame({"Row": range(500)})
    frame.to_parquet(path, index=False)
    return path


class TestRead:
    def test_reads_xlsx(self, sample_xlsx):
        result = run("read.py", str(sample_xlsx))
        assert result.returncode == 0
        assert "Alice" in result.stdout

    def test_reads_xlsx_json(self, sample_xlsx):
        result = run("read.py", str(sample_xlsx), "--json")
        assert result.returncode == 0
        data = json.loads(result.stdout)
        assert "Data" in data
        assert data["Data"][0] == ["Name", "Age", "Score"]

    def test_reads_csv(self, sample_csv):
        result = run("read.py", str(sample_csv))
        assert result.returncode == 0
        assert "Alice" in result.stdout

    def test_reads_parquet(self, sample_parquet):
        result = run("read.py", str(sample_parquet))
        assert result.returncode == 0
        assert "Alice" in result.stdout

    def test_reads_parquet_json(self, sample_parquet):
        result = run("read.py", str(sample_parquet), "--json")
        assert result.returncode == 0
        data = json.loads(result.stdout)
        assert data["Sheet1"][0] == ["Name", "Age", "Score", "Region"]
        # Parquet carries types, so the values arrive as numbers, not strings.
        assert data["Sheet1"][1] == ["Alice", 30, 95, "East"]

    def test_previews_parquet_without_reading_every_row(self, large_parquet):
        result = run("read.py", str(large_parquet), "--limit", "5")

        assert result.returncode == 0
        lines = result.stdout.splitlines()
        assert lines[:2] == ["Row", "0"]
        # The footer row count reports the remainder that the preview skipped.
        assert lines[-1] == "... (496 more rows)"

    def test_parquet_json_emits_every_row(self, large_parquet):
        result = run("read.py", str(large_parquet), "--limit", "5", "--json")

        assert result.returncode == 0
        assert len(json.loads(result.stdout)["Sheet1"]) == 501

    def test_points_numbers_files_at_the_bridge(self, tmp_path):
        source = tmp_path / "deck.numbers"
        source.write_bytes(b"PK\x03\x04")

        result = run("read.py", str(source))

        assert result.returncode != 0
        assert "numbers-bridge.ts" in result.stderr


class TestCreate:
    def test_creates_from_json(self, tmp_path):
        out = tmp_path / "output.xlsx"
        data = json.dumps([{"Name": "Alice", "Score": 95}, {"Name": "Bob", "Score": 82}])
        result = run("create.py", "--output", str(out), "--json", data)
        assert result.returncode == 0
        assert out.exists()
        wb = openpyxl.load_workbook(str(out))
        ws = wb.active
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Alice"

    def test_creates_from_csv(self, sample_csv, tmp_path):
        out = tmp_path / "from_csv.xlsx"
        result = run("create.py", "--output", str(out), "--input", str(sample_csv))
        assert result.returncode == 0
        assert out.exists()

    def test_stores_formula_like_values_as_text(self, tmp_path):
        out = tmp_path / "literal-value.xlsx"
        result = run("create.py", "--output", str(out), "--json", '[["Value"],["=1+1"]]')

        assert result.returncode == 0
        cell = openpyxl.load_workbook(out).active["A2"]
        assert cell.value == "=1+1"
        assert cell.data_type == "s"


class TestLibraryRecipe:
    def test_creates_a_styled_multi_sheet_workbook(self, tmp_path):
        from datetime import date

        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.worksheet.table import Table, TableStyleInfo

        output = tmp_path / "sales.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Sales"
        sheet.append(["Date", "Region", "Revenue", "Target", "Variance"])
        for values in [
            (date(2026, 1, 31), "East", 125000, 120000),
            (date(2026, 1, 31), "West", 117500, 115000),
        ]:
            sheet.append(values)
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, 5, f"=C{row}-D{row}")
            sheet.cell(row, 1).number_format = "mmm d, yyyy"
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
        table = Table(displayName="SalesTable", ref=f"A1:E{sheet.max_row}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
        sheet.freeze_panes = "A2"
        summary = workbook.create_sheet("Summary")
        summary["A1"] = "Total revenue"
        summary["B1"] = "=SUM(Sales!C2:C3)"
        workbook.save(output)

        check = load_workbook(output, data_only=False)
        assert check.sheetnames == ["Sales", "Summary"]
        assert check["Sales"]["E2"].value == "=C2-D2"
        assert check["Sales"]["A2"].number_format == "mmm d, yyyy"
        assert check["Sales"]["A1"].font.bold
        assert check["Sales"].freeze_panes == "A2"
        assert "SalesTable" in check["Sales"].tables
        assert check["Summary"]["B1"].value == "=SUM(Sales!C2:C3)"


class TestParquetRecipe:
    def test_inspects_and_projects_a_parquet_file(self, sample_parquet, tmp_path):
        import pyarrow.parquet as pq

        parquet_file = pq.ParquetFile(sample_parquet)
        assert parquet_file.metadata.num_rows == 3
        assert parquet_file.schema_arrow.names == ["Name", "Age", "Score", "Region"]

        frame = pandas.read_parquet(sample_parquet, columns=["Name", "Region"])
        output = tmp_path / "by-region.parquet"
        frame.to_parquet(output, index=False)

        written = pandas.read_parquet(output)
        assert list(written.columns) == ["Name", "Region"]
        # Parquet keeps column types, including dictionary-encoded categories.
        assert list(written["Region"].cat.categories) == ["East", "West"]


class TestQuery:
    def test_filters_rows(self, sample_xlsx):
        result = run("query.py", str(sample_xlsx), "--filter", "Age > 28")
        assert result.returncode == 0
        assert "Alice" in result.stdout
        assert "Carol" in result.stdout
        assert "Bob" not in result.stdout

    def test_describe(self, sample_xlsx):
        result = run("query.py", str(sample_xlsx), "--describe")
        assert result.returncode == 0
        assert "Age" in result.stdout

    def test_json_output(self, sample_xlsx):
        result = run("query.py", str(sample_xlsx), "--json")
        assert result.returncode == 0
        data = json.loads(result.stdout)
        assert isinstance(data, list)
        assert len(data) == 3

    def test_filters_parquet_rows(self, sample_parquet):
        result = run("query.py", str(sample_parquet), "--filter", "Age > 28")
        assert result.returncode == 0
        assert "Alice" in result.stdout
        assert "Carol" in result.stdout
        assert "Bob" not in result.stdout

    def test_parquet_is_not_parsed_as_csv(self, sample_parquet):
        """A Parquet file decodes to its real records rather than garbage CSV rows."""
        result = run("query.py", str(sample_parquet), "--json")
        assert result.returncode == 0
        assert json.loads(result.stdout) == [
            {"Name": "Alice", "Age": 30, "Score": 95, "Region": "East"},
            {"Name": "Bob", "Age": 25, "Score": 82, "Region": "West"},
            {"Name": "Carol", "Age": 35, "Score": 91, "Region": "East"},
        ]

    def test_saves_result_as_parquet(self, sample_xlsx, tmp_path):
        out = tmp_path / "result.parquet"
        result = run("query.py", str(sample_xlsx), "--filter", "Age > 28", "--output", str(out))
        assert result.returncode == 0
        frame = pandas.read_parquet(out)
        assert list(frame["Name"]) == ["Alice", "Carol"]

    def test_rejects_an_unsupported_input_format(self, tmp_path):
        source = tmp_path / "data.dat"
        source.write_text("Name,Age\nAlice,30\n")

        result = run("query.py", str(source))

        assert result.returncode != 0
        assert "Unsupported format: .dat" in result.stderr

    @pytest.mark.parametrize("suffix", [".numbers", ".xls"])
    def test_points_bridge_formats_at_the_bridge(self, tmp_path, suffix):
        """pandas cannot open these, so the message names the bridge, not pandas."""
        source = tmp_path / f"legacy{suffix}"
        source.write_bytes(b"PK\x03\x04")

        result = run("query.py", str(source))

        assert result.returncode != 0
        assert "numbers-bridge.ts" in result.stderr
        assert "pandas" not in result.stderr


class TestConvert:
    def test_csv_to_xlsx(self, sample_csv, tmp_path):
        out = tmp_path / "converted.xlsx"
        result = run("convert.py", str(sample_csv), "--output", str(out))
        assert result.returncode == 0
        assert out.exists()

    def test_xlsx_to_csv(self, sample_xlsx, tmp_path):
        out = tmp_path / "exported.csv"
        result = run("convert.py", str(sample_xlsx), "--output", str(out))
        assert result.returncode == 0
        content = out.read_text()
        assert "Alice" in content

    def test_stores_formula_like_values_as_text(self, tmp_path):
        source = tmp_path / "formula.csv"
        source.write_text("Value\n=1+1\n")
        out = tmp_path / "literal-value.xlsx"

        result = run("convert.py", str(source), "--output", str(out))

        assert result.returncode == 0
        cell = openpyxl.load_workbook(out).active["A2"]
        assert cell.value == "=1+1"
        assert cell.data_type == "s"

    @pytest.mark.parametrize("suffix", [".csv", ".tsv", ".xlsx"])
    def test_parquet_to_other_formats(self, sample_parquet, tmp_path, suffix):
        out = tmp_path / f"from-parquet{suffix}"

        result = run("convert.py", str(sample_parquet), "--output", str(out))

        assert result.returncode == 0
        if suffix == ".xlsx":
            assert openpyxl.load_workbook(out).active["A2"].value == "Alice"
        else:
            assert "Alice" in out.read_text()

    @pytest.mark.parametrize("source_fixture", ["sample_csv", "sample_xlsx", "sample_parquet"])
    def test_other_formats_to_parquet(self, request, tmp_path, source_fixture):
        source = request.getfixturevalue(source_fixture)
        out = tmp_path / "converted.parquet"

        result = run("convert.py", str(source), "--output", str(out))

        assert result.returncode == 0
        frame = pandas.read_parquet(out)
        assert list(frame["Name"]) == ["Alice", "Bob", "Carol"]
        assert list(frame["Age"]) == [30, 25, 35]

    @pytest.mark.parametrize(
        ("source_suffix", "output_suffix"), [(".dat", ".csv"), (".csv", ".dat")]
    )
    def test_rejects_unsupported_formats(self, tmp_path, source_suffix, output_suffix):
        source = tmp_path / f"data{source_suffix}"
        source.write_text("Name,Age\nAlice,30\n")

        result = run("convert.py", str(source), "--output", str(tmp_path / f"out{output_suffix}"))

        assert result.returncode != 0
        assert "Unsupported format: .dat" in result.stderr

    @pytest.mark.parametrize("suffix", [".numbers", ".xls"])
    def test_points_bridge_formats_at_the_bridge(self, sample_csv, tmp_path, suffix):
        """The bridge handles these in both directions, so neither side suggests pandas."""
        source = tmp_path / f"legacy{suffix}"
        source.write_bytes(b"PK\x03\x04")

        reading = run("convert.py", str(source), "--output", str(tmp_path / "out.csv"))
        writing = run("convert.py", str(sample_csv), "--output", str(tmp_path / f"out{suffix}"))

        for result in (reading, writing):
            assert result.returncode != 0
            assert "numbers-bridge.ts" in result.stderr


class TestEdit:
    def test_stores_formula_like_values_as_text(self, sample_xlsx, tmp_path):
        out = tmp_path / "literal-value.xlsx"
        result = run(
            "edit.py",
            str(sample_xlsx),
            "--set-cell",
            "D2==1+1",
            "--add-row",
            '["=2+2"]',
            "--output",
            str(out),
        )

        assert result.returncode == 0
        sheet = openpyxl.load_workbook(out).active
        assert sheet["D2"].value == "=1+1"
        assert sheet["D2"].data_type == "s"
        assert sheet["A5"].value == "=2+2"
        assert sheet["A5"].data_type == "s"


# A workbook as Excel saves it: a frozen corner plus one selection per pane.
EXCEL_AUTHORED_VIEWS = (
    '<sheetViews><sheetView tabSelected="1" workbookViewId="0">'
    '<pane xSplit="1" ySplit="1" topLeftCell="B2" activePane="bottomRight" state="frozen"/>'
    '<selection pane="topRight" activeCell="B1" sqref="B1"/>'
    '<selection pane="bottomLeft" activeCell="A2" sqref="A2"/>'
    '<selection pane="bottomRight" activeCell="B2" sqref="B2"/>'
    "</sheetView></sheetViews>"
)


def rewrite_sheet_views(source: Path, target: Path, views: str) -> Path:
    """Replace sheet1's sheetViews so the fixture is not openpyxl's own output."""
    with zipfile.ZipFile(source) as archive_in, zipfile.ZipFile(
        target, "w", zipfile.ZIP_DEFLATED
    ) as archive_out:
        for item in archive_in.infolist():
            data = archive_in.read(item.filename)
            if item.filename == "xl/worksheets/sheet1.xml":
                text = data.decode()
                start = text.find("<sheetViews")
                end = text.find("</sheetViews>") + len("</sheetViews>")
                text = text[:start] + views + text[end:]
                data = text.encode()
            archive_out.writestr(item, data)
    return target


def selection_panes(path: Path) -> list[str]:
    namespace = "{http://schemas.openxmlformats.org/spreadsheetml/2006/main}"
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("xl/worksheets/sheet1.xml"))
    view = root.find(f"{namespace}sheetViews/{namespace}sheetView")
    return [s.get("pane") or "topLeft" for s in view.findall(f"{namespace}selection")]


@pytest.fixture
def frozen_source(sample_xlsx, tmp_path) -> Path:
    return rewrite_sheet_views(sample_xlsx, tmp_path / "frozen.xlsx", EXCEL_AUTHORED_VIEWS)


class TestValidate:
    def test_accepts_a_workbook_built_from_scratch(self, sample_xlsx):
        result = run("validate.py", str(sample_xlsx))
        assert result.returncode == 0
        assert "valid" in result.stdout

    def test_accepts_an_excel_authored_workbook(self, frozen_source):
        result = run("validate.py", str(frozen_source))
        assert result.returncode == 0

    @pytest.mark.parametrize(
        ("freeze", "expected"),
        [
            ("A2", 'selections for pane "bottomLeft"'),
            ("B2", "the schema allows at most 4"),
        ],
    )
    def test_rejects_selections_left_behind_by_freeze_panes(
        self, frozen_source, tmp_path, freeze, expected
    ):
        out = tmp_path / f"refrozen-{freeze}.xlsx"
        workbook = openpyxl.load_workbook(frozen_source)
        workbook.active.freeze_panes = freeze
        workbook.save(out)

        result = run("validate.py", str(out))

        assert result.returncode == 1
        assert expected in result.stdout

    def test_fix_repairs_the_sheet_view_and_keeps_the_freeze(self, frozen_source, tmp_path):
        out = tmp_path / "refrozen.xlsx"
        workbook = openpyxl.load_workbook(frozen_source)
        workbook.active.freeze_panes = "B2"
        workbook.save(out)

        assert run("validate.py", str(out), "--fix").returncode == 0
        assert run("validate.py", str(out)).returncode == 0
        assert selection_panes(out) == ["topRight", "bottomLeft", "bottomRight"]

        sheet = openpyxl.load_workbook(out).active
        assert sheet.freeze_panes == "B2"
        assert sheet["A2"].value == "Alice"

    def test_resetting_the_view_avoids_the_problem(self, frozen_source, tmp_path):
        out = tmp_path / "reset.xlsx"
        workbook = openpyxl.load_workbook(frozen_source)
        sheet = workbook.active
        sheet.sheet_view.selection = [Selection()]
        sheet.freeze_panes = "B2"
        workbook.save(out)

        assert run("validate.py", str(out)).returncode == 0

    def test_rejects_a_color_scale_without_colors(self, tmp_path):
        out = tmp_path / "colorscale.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in [["n"], [1], [2]]:
            sheet.append(row)
        sheet.conditional_formatting.add("A2:A3", ColorScaleRule(start_type="min", end_type="max"))
        workbook.save(out)

        result = run("validate.py", str(out))

        assert result.returncode == 1
        assert "colorScale" in result.stdout
