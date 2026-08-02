#!/usr/bin/env python3
"""Convert between spreadsheet formats: CSV, TSV, XLSX, and Parquet."""

import argparse
import sys
from pathlib import Path

FORMATS = ".csv, .tsv, .xlsx, or .parquet"


def unsupported(ext: str) -> str:
    return (
        f"Unsupported format: {ext or '(no extension)'}. Expected {FORMATS}. "
        "Handle any other format with pandas directly."
    )


def store_formula_like_values_as_text(output: str, dataframe):
    from openpyxl import load_workbook

    workbook = load_workbook(output)
    sheet = workbook.active
    for row_index, row in enumerate(dataframe.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            if isinstance(value, str) and value.startswith("="):
                cell = sheet.cell(row=row_index, column=column_index)
                cell.value = value
                cell.data_type = "s"
    workbook.save(output)


def main():
    parser = argparse.ArgumentParser(description="Convert spreadsheet formats")
    parser.add_argument("input", help="Input file (.csv, .tsv, .xlsx, .xlsm, .parquet)")
    parser.add_argument("--output", required=True, help="Output file (.csv, .tsv, .xlsx, .parquet)")
    parser.add_argument("--sheet", help="Source sheet name (for multi-sheet XLSX input)")
    args = parser.parse_args()

    src = Path(args.input).suffix.lower()
    dst = Path(args.output).suffix.lower()

    try:
        import pandas as pd
    except ImportError:
        sys.exit(
            "pandas is unavailable. Reload this skill to retry dependency setup."
        )

    if src in (".xlsx", ".xlsm"):
        df = pd.read_excel(args.input, sheet_name=args.sheet or 0)
    elif src == ".xls":
        sys.exit(
            "Legacy .xls files need the TypeScript compatibility bridge: "
            "tsx scripts/numbers-bridge.ts input.xls --output output.xlsx"
        )
    elif src == ".parquet":
        df = pd.read_parquet(args.input)
    elif src == ".tsv":
        df = pd.read_csv(args.input, sep="\t")
    elif src == ".csv":
        df = pd.read_csv(args.input)
    else:
        # Guessing CSV for an unrecognized extension parses a binary format into
        # garbage rows instead of reporting that it is unsupported.
        sys.exit(unsupported(src))

    if dst == ".xlsx":
        df.to_excel(args.output, index=False)
        store_formula_like_values_as_text(args.output, df)
    elif dst == ".parquet":
        df.to_parquet(args.output, index=False)
    elif dst == ".tsv":
        df.to_csv(args.output, sep="\t", index=False)
    elif dst == ".csv":
        df.to_csv(args.output, index=False)
    else:
        sys.exit(unsupported(dst))

    print(f"Converted {len(df)} rows: {args.input} -> {args.output}")


if __name__ == "__main__":
    main()
