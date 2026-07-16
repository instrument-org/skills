#!/usr/bin/env python3
"""Create a new Excel spreadsheet from JSON data or a CSV file.

Input JSON format:
  A list of objects (rows) or a list of lists.
  If a list of objects, the keys of the first object become the header row.

Examples:
  python scripts/create.py --output report.xlsx --json '[{"Name":"Alice","Age":30}]'
  python scripts/create.py --output report.xlsx --input data.csv
"""

import argparse
import csv
import json
import sys
from pathlib import Path


def set_literal_value(cell, value):
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"


def main():
    parser = argparse.ArgumentParser(description="Create an Excel spreadsheet")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--json", dest="json_data", help="JSON array of rows")
    parser.add_argument("--input", help="Input CSV or TSV file")
    parser.add_argument("--sheet", default="Sheet1", help="Sheet name")
    parser.add_argument("--title")
    parser.add_argument("--freeze-header", action="store_true",
                        help="Freeze the first (header) row")
    args = parser.parse_args()

    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        sys.exit(
            "openpyxl is unavailable. Reload this skill to retry dependency setup."
        )

    # Load data
    if args.json_data:
        raw = json.loads(args.json_data)
        if raw and isinstance(raw[0], dict):
            headers = list(raw[0].keys())
            rows = [[r.get(h) for h in headers] for r in raw]
            rows.insert(0, headers)
        else:
            rows = raw
    elif args.input:
        ext = Path(args.input).suffix.lower()
        delim = "\t" if ext == ".tsv" else ","
        with open(args.input, newline="", encoding="utf-8-sig") as f:
            rows = list(csv.reader(f, delimiter=delim))
    else:
        sys.exit("Provide --json or --input")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = args.sheet

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            set_literal_value(cell, value)
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

    # Auto-fit column widths (approximate)
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    if args.freeze_header:
        ws.freeze_panes = "A2"

    if args.title:
        wb.properties.title = args.title

    wb.save(args.output)
    print(f"Created: {args.output} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
