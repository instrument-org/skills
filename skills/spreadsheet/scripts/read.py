#!/usr/bin/env python3
"""Read and display spreadsheet data (XLSX, XLSM, CSV, TSV, Parquet)."""

import argparse
import csv
import json
import sys
from pathlib import Path

FORMATS = ".xlsx, .xlsm, .csv, .tsv, or .parquet"

BRIDGE = (
    "Apple Numbers and legacy .xls files need the TypeScript compatibility bridge: "
    "tsx scripts/numbers-bridge.ts <input> --output <output>"
)

PYARROW_MISSING = (
    "pyarrow is unavailable, so this run cannot handle Parquet. Reload this skill "
    "to retry dependency setup; Windows on ARM has no pyarrow build."
)


def unsupported(ext: str) -> str:
    if ext in (".numbers", ".xls"):
        return BRIDGE
    return (
        f"Unsupported format: {ext or '(no extension)'}. Expected {FORMATS}. "
        "Handle any other format with pandas directly."
    )


def read_csv(path: str, delimiter: str = ",") -> list[list]:
    with open(path, newline="", encoding="utf-8-sig") as f:
        return list(csv.reader(f, delimiter=delimiter))


def read_parquet(path: str, limit: int | None) -> tuple[list[list], int]:
    """Return the header plus at most `limit` rows, and the row count including the header."""
    try:
        import pyarrow.parquet as pq
    except ImportError:
        sys.exit(PYARROW_MISSING)

    parquet_file = pq.ParquetFile(path)
    names = parquet_file.schema_arrow.names
    rows: list[list] = []
    # Reading in batches keeps a preview from materializing a file that may be
    # many gigabytes; the row count comes from the footer rather than a scan.
    for batch in parquet_file.iter_batches(batch_size=1024 if limit is None else max(limit, 1)):
        # to_pylist() decodes dictionary and nested columns into native Python
        # values, so the rows print and serialize like the other formats do.
        rows.extend([record[name] for name in names] for record in batch.to_pylist())
        if limit is not None and len(rows) >= limit:
            break
    return [names, *rows], parquet_file.metadata.num_rows + 1


def read_excel(path: str, sheet: str | None = None) -> dict[str, list[list]]:
    try:
        import openpyxl
    except ImportError:
        sys.exit(
            "openpyxl is unavailable. Reload this skill to retry dependency setup."
        )

    wb = openpyxl.load_workbook(path, data_only=True)
    names = [sheet] if sheet else wb.sheetnames
    result = {}
    for name in names:
        ws = wb[name]
        result[name] = [[cell.value for cell in row] for row in ws.iter_rows()]
    return result


def main():
    parser = argparse.ArgumentParser(description="Read spreadsheet data")
    parser.add_argument("input", help="Input file (.xlsx, .xlsm, .csv, .tsv, .parquet)")
    parser.add_argument("--sheet", help="Sheet name (Excel only)")
    parser.add_argument("--json", action="store_true", dest="as_json")
    parser.add_argument("--limit", type=int, default=50, help="Max rows to display (default: 50)")
    args = parser.parse_args()

    ext = Path(args.input).suffix.lower()
    total = None

    if ext in (".csv",):
        rows = read_csv(args.input)
        data = {"Sheet1": rows}
    elif ext in (".tsv",):
        rows = read_csv(args.input, delimiter="\t")
        data = {"Sheet1": rows}
    elif ext == ".parquet":
        # --json emits every row, so only the display path can stop reading early.
        rows, total = read_parquet(args.input, None if args.as_json else args.limit)
        data = {"Sheet1": rows}
    elif ext in (".xlsx", ".xlsm"):
        data = read_excel(args.input, args.sheet)
    else:
        sys.exit(unsupported(ext))

    if args.as_json:
        print(json.dumps(data, indent=2, default=str))
    else:
        for sheet_name, rows in data.items():
            if len(data) > 1:
                print(f"\n=== {sheet_name} ===")
            for row in rows[:args.limit]:
                print("\t".join(str(v) if v is not None else "" for v in row))
            remaining = (len(rows) if total is None else total) - args.limit
            if remaining > 0:
                print(f"... ({remaining} more rows)")


if __name__ == "__main__":
    main()
