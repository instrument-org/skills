#!/usr/bin/env python3
"""Edit cells, formulas, and rows in an existing Excel spreadsheet.

Operations: set-cell, set-formula, add-row, delete-row.

Examples:
  python scripts/edit.py workbook.xlsx --set-cell "A1=Hello"
  python scripts/edit.py workbook.xlsx --set-formula "C2=SUM(A2:B2)"
  python scripts/edit.py workbook.xlsx --add-row '["Alice", 30, "=B2*1.1"]' --sheet Data
  python scripts/edit.py workbook.xlsx --set-cell "B1=42" --output updated.xlsx
"""

import argparse
import json
import re
import sys


def set_literal_value(cell, value):
    cell.value = value
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"


def cell_ref(ref: str):
    """Parse 'A1' -> (row=1, col=1) or 'Sheet1!A1' -> (sheet, row, col)."""
    if "!" in ref:
        sheet, addr = ref.split("!", 1)
    else:
        sheet = None
        addr = ref
    m = re.match(r"([A-Za-z]+)(\d+)", addr)
    if not m:
        sys.exit(f"Invalid cell reference: {ref}")
    col_str, row_str = m.group(1).upper(), m.group(2)
    col = sum((ord(c) - ord("A") + 1) * (26 ** i) for i, c in enumerate(reversed(col_str)))
    return sheet, int(row_str), col


def main():
    parser = argparse.ArgumentParser(description="Edit an Excel spreadsheet")
    parser.add_argument("input", help="Input .xlsx file")
    parser.add_argument("--output", help="Output path (default: overwrite input)")
    parser.add_argument("--sheet", help="Target sheet name (default: first sheet)")
    parser.add_argument("--set-cell", metavar="REF=VALUE",
                        help="Set cell value, e.g. A1=Hello or Sheet1!B2=42")
    parser.add_argument("--set-formula", metavar="REF=FORMULA",
                        help="Set a formula, e.g. C2==SUM(A2:B2)")
    parser.add_argument("--add-row", metavar="JSON",
                        help="Append a row (JSON array), e.g. '[\"Alice\",30]'")
    parser.add_argument("--delete-row", type=int, metavar="N",
                        help="Delete row N (1-indexed)")
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        sys.exit(
            "openpyxl is unavailable. Reload this skill to retry dependency setup."
        )

    wb = openpyxl.load_workbook(args.input)
    ws = wb[args.sheet] if args.sheet else wb.active

    if args.set_cell:
        ref, value = args.set_cell.split("=", 1)
        sheet_name, row, col = cell_ref(ref)
        target = wb[sheet_name] if sheet_name else ws
        # Try numeric conversion
        try:
            value = int(value)
        except ValueError:
            try:
                value = float(value)
            except ValueError:
                pass
        set_literal_value(target.cell(row=row, column=col), value)

    if args.set_formula:
        ref, formula = args.set_formula.split("=", 1)
        sheet_name, row, col = cell_ref(ref)
        target = wb[sheet_name] if sheet_name else ws
        # Ensure formula starts with =
        if not formula.startswith("="):
            formula = "=" + formula
        target.cell(row=row, column=col, value=formula)

    if args.add_row:
        values = json.loads(args.add_row)
        row = ws.max_row + 1
        for column, value in enumerate(values, start=1):
            set_literal_value(ws.cell(row=row, column=column), value)

    if args.delete_row:
        ws.delete_rows(args.delete_row)

    out = args.output or args.input
    wb.save(out)
    print(f"Saved: {out}")


if __name__ == "__main__":
    main()
